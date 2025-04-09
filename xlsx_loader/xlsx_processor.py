import os
import time
import uuid
import hashlib
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, List

from dotenv import load_dotenv
from loguru import logger
import openpyxl

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO)
logger.add("xlsx_processor.log", rotation="10 MB")


class XlsxProcessor:
    """
    Processes Excel documents and uploads to Pinecone.

    This is a wrapper around the existing Excel parser that adapts it to
    our Pinecone integration requirements.
    """

    def __init__(self):
        # Chunking parameters
        self.chunk_size = int(os.getenv("CHUNK_SIZE", "500"))
        self.chunk_overlap = int(os.getenv("CHUNK_OVERLAP", "75"))

    def process_file(
        self, file_path: str, index=None, namespace: str = None, customer_id: str = None
    ) -> bool:
        """
        Process an XLSX file and store its content in Pinecone.

        Args:
            file_path: Path to the XLSX file
            index: Pinecone index object
            namespace: Pinecone namespace
            customer_id: Customer ID for namespace

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Open workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            # Process workbook
            if not self._process_workbook(
                workbook, file_path, index, namespace, customer_id
            ):
                logger.error("Failed to process workbook")
                return False

            # Close workbook
            workbook.close()

            logger.info(f"Successfully processed XLSX file: {file_path}")
            return True

        except Exception as e:
            logger.error(f"Error processing XLSX file: {str(e)}")
            return False

    def _use_fallback_parser(self, file_path: str) -> None:
        """
        Use a simple fallback parser when the advanced one is not available.

        Args:
            file_path: Path to the Excel file
        """
        try:
            import pandas as pd

            # Base metadata for all chunks from this file
            metadata_base = {
                "filename": Path(file_path).name,
                "file_type": "xlsx",
                "original_path": str(Path(file_path).absolute()),
                "source_type": "local_xlsx",
                "upload_timestamp": datetime.now().isoformat(),
            }

            # Read Excel file using pandas - with try/except to handle different formats
            try:
                excel_file = pd.ExcelFile(file_path)
                sheet_names = excel_file.sheet_names

                for sheet_idx, sheet_name in enumerate(sheet_names):
                    # Read the sheet
                    df = pd.read_excel(file_path, sheet_name=sheet_name)

                    # Skip empty sheets
                    if df.empty:
                        continue

                    # Convert to text representation
                    table_text = []

                    # Get headers
                    headers = df.columns.tolist()

                    # Process each row
                    for _, row in df.iterrows():
                        row_text = []
                        for i, header in enumerate(headers):
                            cell_value = row.iloc[i]
                            if pd.notna(cell_value):  # Skip NaN values
                                row_text.append(f"{header}: {cell_value}")

                        if row_text:
                            table_text.append("; ".join(row_text))

                    # Create chunks
                    if table_text:
                        # Add sheet name if it's not just "Sheet1", "Sheet2", etc.
                        if not sheet_name.lower().startswith("sheet"):
                            for i in range(len(table_text)):
                                table_text[i] += f" —{sheet_name}"

                        text_content = "\n".join(table_text)
                        chunks = self._chunk_text(text_content)

                        # Update metadata with sheet information
                        sheet_metadata = metadata_base.copy()
                        sheet_metadata["content_type"] = "spreadsheet"
                        sheet_metadata["sheet_name"] = sheet_name
                        sheet_metadata["sheet_index"] = sheet_idx

                        # Create and upsert vectors
                        self._create_and_upsert_vectors(chunks, sheet_metadata)

            except Exception as e:
                logger.error(f"Error reading Excel with pandas: {str(e)}")
                # Try CSV as fallback for Excel fallback
                try:
                    df = pd.read_csv(file_path)

                    # Process similar to Excel
                    table_text = []
                    headers = df.columns.tolist()

                    for _, row in df.iterrows():
                        row_text = []
                        for i, header in enumerate(headers):
                            cell_value = row.iloc[i]
                            if pd.notna(cell_value):
                                row_text.append(f"{header}: {cell_value}")

                        if row_text:
                            table_text.append("; ".join(row_text))

                    if table_text:
                        text_content = "\n".join(table_text)
                        chunks = self._chunk_text(text_content)

                        csv_metadata = metadata_base.copy()
                        csv_metadata["content_type"] = "csv"

                        self._create_and_upsert_vectors(chunks, csv_metadata)

                except Exception as csv_error:
                    logger.error(f"Error reading as CSV: {str(csv_error)}")
                    raise

        except Exception as e:
            logger.error(f"Error in fallback parser for {file_path}: {str(e)}")
            raise

    def _process_text_content(self, text_content, file_path: str) -> None:
        """
        Process extracted text content and upload to Pinecone.

        Args:
            text_content: Extracted text content from the Excel file
            file_path: Path to the Excel file
        """
        # Base metadata for all chunks from this file
        metadata_base = {
            "filename": Path(file_path).name,
            "file_type": "xlsx",
            "original_path": str(Path(file_path).absolute()),
            "source_type": "local_xlsx",
            "upload_timestamp": datetime.now().isoformat(),
            "content_type": "spreadsheet_text",
        }

        # Process the text content in batches
        batch_size = 100
        all_chunks = []

        for i in range(0, len(text_content), batch_size):
            batch = text_content[i : i + batch_size]
            batch_text = "\n".join([line for line in batch if line.strip()])

            # Only process if there is content
            if batch_text.strip():
                chunks = self._chunk_text(batch_text)

                # Create metadata for this batch
                batch_metadata = metadata_base.copy()
                batch_metadata["batch_index"] = i // batch_size

                # Accumulate chunks
                all_chunks.extend(chunks)

                # Process if we have enough chunks
                if len(all_chunks) >= 50:
                    self._create_and_upsert_vectors(all_chunks, batch_metadata)
                    all_chunks = []

        # Process any remaining chunks
        if all_chunks:
            batch_metadata = metadata_base.copy()
            batch_metadata["batch_index"] = len(text_content) // batch_size
            self._create_and_upsert_vectors(all_chunks, batch_metadata)

    def _process_html_tables(self, html_tables, file_path: str) -> None:
        """
        Process HTML table representations of Excel data and upload to Pinecone.

        Args:
            html_tables: HTML table content from the Excel file
            file_path: Path to the Excel file
        """
        # Base metadata for all chunks from this file
        metadata_base = {
            "filename": Path(file_path).name,
            "file_type": "xlsx",
            "original_path": str(Path(file_path).absolute()),
            "source_type": "local_xlsx_html",
            "upload_timestamp": datetime.now().isoformat(),
            "content_type": "spreadsheet_html",
        }

        # Process each HTML table
        for i, table_html in enumerate(html_tables):
            if not table_html.strip():
                continue

            # Extract table caption (sheet name)
            sheet_name = ""
            import re

            caption_match = re.search(r"<caption>(.*?)</caption>", table_html)
            if caption_match:
                sheet_name = caption_match.group(1)

            # Chunk the HTML content - this preserves the table structure
            # but may not be semantically ideal
            chunks = self._chunk_text(table_html)

            # Create metadata for this table
            table_metadata = metadata_base.copy()
            table_metadata["table_index"] = i
            if sheet_name:
                table_metadata["sheet_name"] = sheet_name

            # Create and upsert vectors
            self._create_and_upsert_vectors(chunks, table_metadata)

    def _chunk_text(self, text: str) -> List[str]:
        """
        Split text into chunks based on configuration.

        Args:
            text: The text to chunk

        Returns:
            List of text chunks
        """
        if not text or len(text.strip()) == 0:
            return []

        chunks = []
        start = 0
        text_length = len(text)

        while start < text_length:
            # Calculate end position with overlap
            end = min(start + self.chunk_size, text_length)

            # Get the chunk
            chunk = text[start:end]

            # Only add non-empty chunks
            if chunk.strip():
                chunks.append(chunk)

            # Move start position for next chunk, considering overlap
            start = end - self.chunk_overlap if end < text_length else text_length

        return chunks

    def _generate_embedding(self, text: str) -> List[float]:
        """
        Generate embedding for text using Pinecone's server-side embedding.

        Args:
            text: Text to embed

        Returns:
            Vector embedding
        """
        try:
            # Use server-side embedding by passing the text directly
            # The embedding will be generated by Pinecone
            return text
        except Exception as e:
            logger.error(f"Error preparing text for server-side embedding: {str(e)}")
            raise

    def _generate_vector_id(self, text: str, metadata: Dict[str, Any]) -> str:
        """
        Generate a unique ID for a vector.

        Args:
            text: The text content
            metadata: Metadata for the vector

        Returns:
            Unique vector ID
        """
        # Create a unique hash based on content and metadata
        content_to_hash = text

        # Add key metadata to the hash
        filename = metadata.get("filename", "")
        if filename:
            content_to_hash += filename

        chunk_index = metadata.get("chunk_index", 0)
        content_to_hash += str(chunk_index)

        # Generate a hash
        content_hash = hashlib.md5(content_to_hash.encode("utf-8")).hexdigest()

        # Add timestamp for uniqueness
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")

        # Combine with a UUID component
        unique_id = f"file_{content_hash}_{timestamp}_{str(uuid.uuid4())[:8]}"

        return unique_id

    def _create_and_upsert_vectors(
        self,
        chunks: List[str],
        index=None,
        namespace: str = None,
        customer_id: str = None,
    ) -> bool:
        """
        Create vectors from chunks and upsert to Pinecone with improved metadata.

        Args:
            chunks: List of text chunks
            index: Pinecone index object
            namespace: Pinecone namespace
            customer_id: Customer ID for namespace

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            vectors = []
            timestamp = datetime.now().isoformat()

            for i, chunk in enumerate(chunks):
                # Create enhanced metadata for each chunk
                metadata = {
                    "text": chunk,
                    "chunk_index": i,
                    "total_chunks": len(chunks),
                    "upload_timestamp": timestamp,
                    "source_type": "local_xlsx",
                }

                # Generate a unique ID for the vector
                vector_id = f"xlsx_{i}_{hashlib.md5(chunk.encode()).hexdigest()[:8]}_{int(time.time())}"

                # Create vector with properly formatted values for server-side embedding
                vector = {
                    "id": vector_id,
                    "values": [0.1]
                    * 1024,  # Placeholder values for server-side embedding
                    "metadata": metadata,
                }
                vectors.append(vector)

            # Upsert vectors to Pinecone
            if index and vectors:
                # Use customer-specific namespace if provided
                effective_namespace = namespace
                if customer_id and namespace:
                    effective_namespace = f"{namespace}_{customer_id}"

                # Use batch processing for better performance
                batch_size = 100
                for i in range(0, len(vectors), batch_size):
                    batch = vectors[i : i + batch_size]
                    index.upsert(vectors=batch, namespace=effective_namespace)
                    logger.info(f"Upserted batch of {len(batch)} vectors to Pinecone")

                logger.info(
                    f"Successfully upserted total of {len(vectors)} vectors to Pinecone"
                )
            else:
                if not index:
                    logger.warning("No Pinecone index provided, skipping upsert")
                if not vectors:
                    logger.warning("No vectors to upsert")

            return True

        except Exception as e:
            logger.error(f"Error upserting vectors: {str(e)}")
            return False

    def close(self):
        """Close any open file handles."""
        if hasattr(self, "workbook") and self.workbook is not None:
            self.workbook.close()
            self.workbook = None

    def _process_workbook(
        self,
        workbook,
        file_path: str,
        index=None,
        namespace: str = None,
        customer_id: str = None,
    ) -> bool:
        """Process the workbook and extract content in a structured format."""
        try:
            # Process each sheet
            for sheet_idx, sheet in enumerate(workbook.worksheets):
                # Skip empty sheets
                if sheet.max_row <= 1 and sheet.max_column <= 1:
                    continue

                # Get sheet name
                sheet_name = sheet.title

                # Process sheet content as structured data
                all_texts = self._extract_structured_sheet_data(sheet)

                if not all_texts:
                    continue

                # Create chunks
                chunks = self._chunk_text("\n\n".join(all_texts))

                # Create and upsert vectors
                self._create_and_upsert_vectors(chunks, index, namespace, customer_id)

            return True

        except Exception as e:
            logger.error(f"Error processing workbook: {str(e)}")
            return False

    def _extract_structured_sheet_data(self, sheet):
        """
        Extract data from a sheet in a structured format that preserves table relationships.

        Args:
            sheet: The Excel worksheet

        Returns:
            List of formatted text blocks
        """
        # Skip if the sheet is too small
        if sheet.max_row <= 1 or sheet.max_column <= 1:
            return []

        # Get all data including headers
        all_data = []
        for row in sheet.iter_rows():
            row_data = []
            for cell in row:
                # Convert cell value to string, handle None
                value = str(cell.value) if cell.value is not None else ""
                row_data.append(value.strip())
            # Only include rows that have at least one non-empty cell
            if any(cell for cell in row_data):
                all_data.append(row_data)

        if not all_data:
            return []

        # Try to identify if first row is headers
        headers = all_data[0] if all_data else []
        has_headers = False

        # Check if headers are meaningful (not all empty/numeric)
        if headers and any(h and not h.isdigit() for h in headers):
            has_headers = True
            data_rows = all_data[1:]
        else:
            data_rows = all_data

        # Format as structured text
        result_texts = []

        # Option 1: If we have headers, format as key-value pairs
        if has_headers and data_rows:
            for row in data_rows:
                # Skip rows with all empty values
                if not any(cell for cell in row):
                    continue

                row_parts = []
                for i, value in enumerate(row):
                    if i < len(headers) and headers[i] and value:
                        row_parts.append(f"{headers[i].strip()}: {value.strip()}")

                if row_parts:
                    result_texts.append(" | ".join(row_parts))

        # Option 2: Format as a text table
        else:
            # Calculate column widths for proper alignment
            col_widths = []
            for row in all_data:
                while len(col_widths) < len(row):
                    col_widths.append(0)
                for i, cell in enumerate(row):
                    col_widths[i] = max(col_widths[i], len(str(cell)))

            # Format as ASCII table
            for row in all_data:
                formatted_row = []
                for i, cell in enumerate(row):
                    if i < len(col_widths) and cell:
                        formatted_row.append(cell)
                if formatted_row:
                    result_texts.append(" | ".join(formatted_row))

        # Add sheet name as context
        sheet_context = f"Sheet: {sheet.title}"
        result_texts.insert(0, sheet_context)

        return result_texts

    def dry_run(self, file_path: str, output_dir: str = "dry_run_output") -> bool:
        """
        Process the file without uploading to Pinecone, saving chunks as structured markdown files.

        Args:
            file_path: Path to the XLSX file
            output_dir: Directory to save the markdown files

        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Create output directory if it doesn't exist
            output_path = Path(output_dir)
            output_path.mkdir(exist_ok=True)

            # Open workbook
            workbook = openpyxl.load_workbook(file_path, data_only=True)

            # Get file information
            file_name = Path(file_path).name
            file_stem = Path(file_path).stem
            file_path_abs = str(Path(file_path).absolute())

            # Common timestamp for all chunks in this file
            timestamp = datetime.now().isoformat()

            # Process each sheet
            for sheet in workbook.worksheets:
                # Skip empty sheets
                if sheet.max_row <= 1 and sheet.max_column <= 1:
                    continue

                sheet_name = sheet.title
                logger.info(f"Processing sheet: {sheet_name}")

                # Extract sheet data with better structure preservation
                all_texts = self._extract_structured_sheet_data(sheet)

                if not all_texts:
                    logger.warning(f"No usable content in sheet {sheet_name}")
                    continue

                # Create a single string representing the formatted sheet content
                sheet_text = "\n\n".join(all_texts)

                # Create chunks from text content
                chunks = self._chunk_text(sheet_text)

                if not chunks:
                    logger.warning(f"No chunks created from sheet {sheet_name}")
                    continue

                # Save chunks to markdown files
                for i, chunk in enumerate(chunks):
                    # Create comprehensive metadata for each chunk (without duplicating text)
                    metadata = {
                        "chunk_index": i,
                        "total_chunks": len(chunks),
                        "filename": file_name,
                        "file_type": "xlsx",
                        "original_path": file_path_abs,
                        "source_type": "local_xlsx",
                        "sheet_name": sheet_name,
                        "upload_timestamp": timestamp,
                        "chunk_size": self.chunk_size,
                        "chunk_overlap": self.chunk_overlap,
                        "chunk_length_chars": len(chunk),
                        "chunk_length_tokens": len(chunk.split()),
                    }

                    # Create filename with sheet name and chunk number in consistent format
                    filename = (
                        f"{file_stem}_{sheet_name}_chunk_{i+1}_of_{len(chunks)}.md"
                    )
                    chunk_path = output_path / filename

                    # Format chunk for better readability - always use code blocks for Excel data
                    formatted_chunk = f"```\n{chunk}\n```"

                    # Write chunk to file with improved formatting
                    with open(chunk_path, "w", encoding="utf-8") as f:
                        f.write(
                            f"# {file_name} - {sheet_name} - Chunk {i+1} of {len(chunks)}\n\n"
                        )
                        f.write("## Content\n\n")
                        f.write(formatted_chunk)
                        f.write("\n\n## Metadata\n\n")

                        # Write metadata in a clean, organized format
                        for key, value in sorted(metadata.items()):
                            f.write(f"- **{key}**: {value}\n")

            # Close workbook
            workbook.close()

            logger.info(f"Dry run completed. Chunks saved to {output_dir}")
            return True

        except Exception as e:
            logger.error(f"Error in dry run: {str(e)}")
            return False

    def _format_chunk_for_markdown(self, chunk: str) -> str:
        """
        Format a text chunk for better readability in markdown.

        Args:
            chunk: The text chunk to format

        Returns:
            Formatted text suitable for markdown display
        """
        # Ensure the chunk is properly formatted for markdown
        # 1. Replace newlines with double newlines for proper paragraph breaks
        # 2. Escape any markdown special characters if needed

        # Wrap the content in a markdown code block if it contains special characters
        # that might interfere with markdown rendering
        special_chars = ["#", "*", "_", "`", "[", "]", "(", ")", "<", ">", "|"]
        needs_code_block = any(char in chunk for char in special_chars)

        if needs_code_block:
            return f"```\n{chunk}\n```"

        # Otherwise, just ensure paragraphs are properly separated
        formatted = chunk.replace("\n", "\n\n")
        return formatted


def process_xlsx(file_path: str) -> None:
    """
    Process an Excel file and upload to Pinecone.

    Args:
        file_path: Path to the Excel file
    """
    processor = XlsxProcessor()
    processor.process_file(file_path)


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(
        description="Process an Excel file and upload to Pinecone"
    )
    parser.add_argument("file_path", help="Path to the Excel file")

    args = parser.parse_args()
    process_xlsx(args.file_path)
